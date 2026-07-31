# -*- coding: utf-8 -*-
"""
Importa os indicadores gerenciais da planilha "Vivaz - Números.xlsx" para SQL.

- Aba "Dados"  -> escopo 'realizado'  (blocos com DATA na coluna D: 2018..2026)
- Aba "Metas"  -> escopo 'meta'        (blocos com DATA na coluna D: 2019..2026)

Ambas as abas usam o MESMO layout de colunas (a coluna A pode estar
desatualizada; o ano/mês corretos vêm da DATA na coluna D).

Extrai APENAS os inputs canônicos por (ano, mês). Os indicadores derivados
(ocupação, diária média, RevPAR, faturamento, EBITDA, resultado, etc.) são
calculados pelo backend com as mesmas fórmulas da planilha — não são importados.

Gera: sql/21_indicadores_seed.sql  (idempotente, com upsert ON CONFLICT).

Uso:  python scripts/import_indicadores.py
"""
import os
import openpyxl
from openpyxl.utils import column_index_from_string as cidx

HERE = os.path.dirname(os.path.abspath(__file__))
SISTEMA = os.path.dirname(HERE)
XLSX = os.path.normpath(os.path.join(SISTEMA, "..", "importacao", "Vivaz - Números.xlsx"))
OUT = os.path.join(SISTEMA, "sql", "21_indicadores_seed.sql")

# Mapeamento único (layout da aba Metas, também usado pelos blocos-data da Dados).
# (campo_no_banco, coluna_na_planilha)
COLS = [
    ("rn",                   "I"),
    ("receita_hospedagem",   "K"),   # "Receita de Diárias"
    ("pax",                  "Q"),
    ("frigobar",             "U"),
    ("room_service",         "V"),
    ("bar_gaia",             "W"),
    ("rest_allegro",         "X"),
    ("rest_terraza",         "Y"),
    ("map_comercial",        "Z"),
    ("eventos_banquetes",    "AA"),
    ("eventos",              "AC"),
    ("outras_receitas",      "AD"),
    ("nao_operacional",      "AE"),
    ("cmv",                  "AH"),
    ("csp",                  "AI"),
    ("impostos_faturamento", "AJ"),
    ("desp_operacional",     "AK"),
    ("desp_pessoal",         "AL"),
    ("desp_vendas",          "AM"),
    ("pessoal_zz",           "AO"),
    ("despesas_zz",          "AP"),
    ("csll_ir",              "AR"),
    ("investimentos",        "AS"),
    ("map_repasse",          "O"),   # MAP
    ("cafe_repasse",         "P"),   # Café da Manhã
    ("qtd_equipe",           "AV"),
]
FIELDS = [c[0] for c in COLS]


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except ValueError:
        return 0.0


# 2018 usa layout LEGADO (diferente de 2019+):
#  - receita de hospedagem na coluna N (não K)
#  - repasses MAP/Café nas colunas AX/AY (não O/P)
#  - sem CSLL/IR, Investimentos, Pessoal ZZ, Despesas ZZ mensais (ficam 0)
#  - meses identificados por NOME na coluna D (não data)
COLS_2018 = [(f, ("N" if f == "receita_hospedagem" else
                  "AX" if f == "map_repasse" else
                  "AY" if f == "cafe_repasse" else col)) for f, col in COLS]
MESES_NOME = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4, "maio": 5,
    "junho": 6, "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10,
    "novembro": 11, "dezembro": 12,
}


def extract_2018(ws):
    """2018 (layout legado): linhas com nome de mês na col D e ano 2018 na col A."""
    out = {}
    for r in range(1, 30):
        d = ws.cell(r, cidx("D")).value
        a = ws.cell(r, cidx("A")).value
        if not (isinstance(d, str) and d.strip().lower() in MESES_NOME):
            continue
        if a != 2018:
            continue
        rn_val = ws.cell(r, cidx("I")).value
        if not isinstance(rn_val, (int, float)):
            continue
        m = MESES_NOME[d.strip().lower()]
        out[(2018, m)] = {field: num(ws.cell(r, cidx(col)).value) for field, col in COLS_2018}
    return out


def extract(ws):
    """Retorna {(year, month): {field: value}} a partir dos blocos com DATA em col D."""
    out = {}
    dups = []
    for r in range(1, ws.max_row + 1):
        d = ws.cell(r, cidx("D")).value
        if not hasattr(d, "month"):          # só linhas cujo D é data (mês real)
            continue
        rn_val = ws.cell(r, cidx("I")).value  # precisa ter RN numérico
        if not isinstance(rn_val, (int, float)):
            continue
        y, m = d.year, d.month
        row = {field: num(ws.cell(r, cidx(col)).value) for field, col in COLS}
        key = (y, m)
        if key in out:
            dups.append(key)
        out[key] = row
    return out, dups


def sql_val(x):
    return repr(round(float(x), 4))


def emit(escopo, data, f):
    """Um único INSERT multi-linha por escopo (compacto, idempotente)."""
    cols = ", ".join(FIELDS)
    updates = ", ".join(f"{fld} = EXCLUDED.{fld}" for fld in FIELDS)
    tuples = []
    for (y, m) in sorted(data.keys()):
        row = data[(y, m)]
        vals = ", ".join(sql_val(row[fld]) for fld in FIELDS)
        tuples.append(f"('{escopo}', {y}, {m}, {vals})")
    f.write(f"INSERT INTO public.indicadores_mensais (escopo, year, month, {cols})\nVALUES\n")
    f.write(",\n".join(tuples))
    f.write(f"\nON CONFLICT (escopo, year, month) DO UPDATE SET {updates};\n")


def main():
    if not os.path.exists(XLSX):
        raise SystemExit(f"Planilha não encontrada: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    dados, dup_d = extract(wb["Dados"])
    metas, dup_m = extract(wb["Metas"])
    # Realizado: 2018..2025 (2018 vem do bloco legado; 2026+ será digitado no sistema).
    MAX_REALIZADO_YEAR = 2025
    dados = {(y, m): v for (y, m), v in dados.items() if y <= MAX_REALIZADO_YEAR}
    dados.update(extract_2018(wb["Dados"]))  # 2018 com mapeamento legado
    anos = sorted({y for (y, _m) in list(dados) + list(metas)})

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("-- Seed de indicadores (gerado por scripts/import_indicadores.py)\n")
        f.write("-- Execute DEPOIS de sql/20_indicadores.sql. Idempotente (upsert).\n\n")
        f.write("-- Parâmetros por ano (UHs). Ajuste se o nº de unidades mudar por ano.\n")
        for y in anos:
            f.write(
                "INSERT INTO public.indicadores_parametros (year, uhs) "
                f"VALUES ({y}, 172) ON CONFLICT (year) DO NOTHING;\n"
            )
        f.write("\n-- REALIZADO (aba Dados)\n")
        emit("realizado", dados, f)
        f.write("\n-- META (aba Metas)\n")
        emit("meta", metas, f)

    print(f"OK -> {OUT}")
    print(f"  realizado: {len(dados)} meses, anos={sorted({y for y,_ in dados})}")
    print(f"  meta:      {len(metas)} meses, anos={sorted({y for y,_ in metas})}")
    if dup_d:
        print(f"  AVISO duplicatas realizado: {sorted(set(dup_d))}")
    if dup_m:
        print(f"  AVISO duplicatas meta: {sorted(set(dup_m))}")


if __name__ == "__main__":
    main()
